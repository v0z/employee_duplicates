#!/usr/bin/env python3

import random

import openpyxl
from openpyxl.styles import PatternFill

from helper import *

# constants

"""
mode 1 - default (finding and marking duplicate emails)
mode 2 - try to juxtapose the name in timetable with an email in performance pivot-table
"""
filename = 'VSP_2021_06_16-30.xlsx'
MODE = 1
WORK_SCHEDULE_SHEET_NAME = 'timetable'  # don't forget to create a tab named timetable in the excel-file
employee_list = {}
result = []
ratio_threshold = 75


def run():
    if not MODE == 2:
        list_of_emails = get_the_first_column_from_excel(filename)

        for index, email in enumerate(list_of_emails, 2):
            if isinstance(email, int):  # int typeerror fix...
                continue
            if '@' not in email:
                continue
            name = get_name_from_email_address(email)
            employee_list[index] = name
            possible_duplicates = compare_names(index, name, employee_list, ratio_threshold)

            if possible_duplicates:
                row_idx = possible_duplicates[0][0]
                username = possible_duplicates[0][1]
                similarity_ratio = possible_duplicates[0][2]

                can_be_marked_as_duplicate = False
                if similarity_ratio > 90:
                    can_be_marked_as_duplicate = True
                else:
                    can_be_marked_as_duplicate = confirm_cell_write(similarity_ratio, index, f'{name}->{username}')

                if can_be_marked_as_duplicate:
                    print(index, name, possible_duplicates)
                    mark_duplicate_rows(index, row_idx)

    # when finished marking the duplicates, try finding the hours for each employee
    # from  the timetable tab
    else:
        """
        Need to prepare the file first:
        add a tab named according to name in const WORK_SCHEDULE_SHEET_NAME. By default - 'timetable'
        Copy the timetable data to the newly created tab as values (No formulas)
        Delete all columns except 'name' and 'total_hours (the column titles don't matter)
        """
        juxtapose_employee_with_timetable()


def juxtapose_employee_with_timetable():
    try:
        email_df = pandas.read_excel(filename, engine='openpyxl', usecols=[0])
        timetable_df = pandas.read_excel(filename, sheet_name=WORK_SCHEDULE_SHEET_NAME, engine='openpyxl')
        name_hours_list = timetable_df.values.tolist()

        # drop the last row
        email_df.drop(email_df.tail(1).index, inplace=True)
        email_df = email_df.values.tolist()
        wb = openpyxl.load_workbook(filename)
        main_sheet = wb.active
        # main_sheet = wb['main']
        timetable_sheet = wb[WORK_SCHEDULE_SHEET_NAME]
        for index, email in enumerate(email_df, 2):
            if isinstance(email, int):  # int typeerror fix...
                continue
            if '@' not in email[0]:  # get email out of list
                continue
            name = get_name_from_email_address(email[0])
            employee_list[index] = name
            match_in_timetable = compare_name_with_timetable(index, name, name_hours_list, ratio_threshold)
            if match_in_timetable:
                row_index = match_in_timetable[0][0]
                timetable_idx = match_in_timetable[0][2]
                username = match_in_timetable[0][3]
                hours = match_in_timetable[0][4]
                ratio = match_in_timetable[0][5]

                can_write_to_cell = False
                if ratio > 90:
                    can_write_to_cell = True
                else:
                    can_write_to_cell = confirm_cell_write(ratio, row_index, f'{name}->{username} | hours: {hours}')

                if can_write_to_cell:
                    column = 8  # column 'H' in excel
                    print(f'ratio: {ratio} | index {row_index} |, names: {name}->{username} | hours: {hours}')
                    write_result = write_data_to_cell(main_sheet=main_sheet, timetable_sheet=timetable_sheet,
                                                      row=row_index, timetable_row=timetable_idx, column=column,
                                                      data=hours)

        wb.save(filename)
    except Exception as exception:
        print('Bit happens - ', exception)


colors = []


def get_new_color():
    color = "%08x" % random.randint(255, 0xBFBFBF)
    if color not in colors:
        colors.append(color)
        return color
    else:
        get_new_color()


def mark_duplicate_rows(row, duplicate_row):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    # for cell in sheet[f'A{sheet.min_row-1}:A{sheet.max_row-1}']:
    color = get_new_color()
    color_fill = PatternFill(start_color=color, fill_type='solid')
    sheet[f'A{row}:A{row}'][0][0].fill = color_fill
    sheet[f'A{duplicate_row}:A{duplicate_row}'][0][0].fill = color_fill
    wb.save(filename)


def write_data_to_cell(main_sheet, timetable_sheet, row, timetable_row, column, data):
    try:
        result = main_sheet.cell(row=row, column=column).value = data

        # mark the processed rows (employees) in the timetable tab
        color_fill = PatternFill(start_color='00FF00', fill_type='solid')
        timetable_sheet.cell(row=timetable_row, column=1).fill = color_fill
        return result
    except Exception as exception:
        raise Exception('Error while writing to cell \n', exception)


if __name__ == '__main__':
    run()
